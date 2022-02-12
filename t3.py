import os
import json
from openpyxl import load_workbook

try:
        if os.path.isfile(r"/home/g/t.xlsx") is True:
         print("OK")
except IOError:
        print("File not available")


workbook = load_workbook(filename="t.xlsx")
sheet = workbook.active

products = {}

# This will run thrugh the columns set to get the data and outputs as tuple
for row in sheet.iter_rows(min_row=3,
                           min_col=1,
                           max_col=12,
                           values_only=True):
    item = row[0]
    columns = {
        "COUNTRY": row[1],
        "NAME OF MANUFACTURERS": row[2],
        "SHIPPER": row[3],
	"ARTICLE": row[4],
	"CONTENTS OF VIOLATION": row[5],
	"QUARANTINE STATION": row[6],
	"NAME OF IMPORTERS": row[7],
	"CAUSE OF VIOLATION": row[8],
	"DISPOSAL OF THE CARGO": row[9],
	"REMARKS": row[10],
#	"PUBLICATION DAY": row[11]
    }
    
    products[item] = columns
#    print(products[item])  prints only the values, not the item 


print(json.dumps(products))

